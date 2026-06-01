import io
import pytest
import pandas as pd
import openpyxl
from excel.editor import aplicar_edicion, exportar_xlsx, EditorError


@pytest.fixture
def df_ventas():
    return pd.DataFrame({
        "Producto":  ["A", "B", "A", "C", "B"],
        "Región":    ["Norte", "Sur", "Norte", "Sur", "Norte"],
        "Precio":    [10.0, 20.0, 10.0, 30.0, 20.0],
        "Cantidad":  [5, 3, 8, 2, 4],
        "Descuento": [0.1, None, 0.2, None, 0.1],
    })


# ── añadir_columna ────────────────────────────────────────────────────────────

def test_añadir_columna_multiplicacion(df_ventas):
    df, desc, extras = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "Total",
        "col1": "Precio", "operador": "*", "col2": "Cantidad",
    })
    assert "Total" in df.columns
    assert df.loc[0, "Total"] == 50.0   # 10 * 5
    assert extras is None


def test_añadir_columna_suma(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "PrecioCant",
        "col1": "Precio", "operador": "+", "col2": "Cantidad",
    })
    assert df.loc[0, "PrecioCant"] == 15.0   # 10 + 5


def test_añadir_columna_valor_fijo(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "PrecioIVA",
        "col1": "Precio", "operador": "*", "valor_fijo": 1.21,
    })
    assert abs(df.loc[0, "PrecioIVA"] - 12.1) < 0.01


def test_añadir_columna_operador_invalido(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {
            "op": "añadir_columna", "nombre": "X",
            "col1": "Precio", "operador": "**", "col2": "Cantidad",
        })


def test_añadir_columna_redondear(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "PrecioRed",
        "col1": "Precio", "operador": "redondear", "valor_fijo": 0,
    })
    assert "PrecioRed" in df.columns
    assert all(df["PrecioRed"] == df["Precio"].round(0))
    assert "redondear" in desc.lower()


def test_añadir_columna_abs(df_ventas):
    df_neg = df_ventas.copy()
    df_neg["Precio"] = [-10, -20, 30, -5, 15]
    df, desc, _ = aplicar_edicion(df_neg, {
        "op": "añadir_columna", "nombre": "PrecioAbs",
        "col1": "Precio", "operador": "abs",
    })
    assert "PrecioAbs" in df.columns
    assert all(df["PrecioAbs"] >= 0)


def test_añadir_columna_raiz(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "RaizPrecio",
        "col1": "Precio", "operador": "raiz",
    })
    assert "RaizPrecio" in df.columns
    import math
    assert abs(float(df["RaizPrecio"].iloc[0]) - math.sqrt(10.0)) < 0.001


def test_añadir_columna_funcion_desconocida_lanza_error(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {
            "op": "añadir_columna", "nombre": "X",
            "col1": "Precio", "operador": "INVENTADA",
        })


def test_añadir_columna_sin_nombre(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {
            "op": "añadir_columna",
            "col1": "Precio", "operador": "*", "col2": "Cantidad",
        })


def test_añadir_columna_expresion_tres_operandos(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "Total",
        "expresion": "Precio * Cantidad",
    })
    assert "Total" in df.columns
    assert df.loc[0, "Total"] == pytest.approx(10 * 5)
    assert "Total" in desc


def test_añadir_columna_expresion_con_valor_fijo(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "añadir_columna", "nombre": "PrecioIVA",
        "expresion": "Precio * 1.21",
    })
    assert "PrecioIVA" in df.columns
    assert df.loc[0, "PrecioIVA"] == pytest.approx(10 * 1.21)


def test_añadir_columna_expresion_invalida_lanza_error(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {
            "op": "añadir_columna", "nombre": "X",
            "expresion": "ColumnaInexistente * 2",
        })


# ── ordenar ───────────────────────────────────────────────────────────────────

def test_ordenar_desc(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {"op": "ordenar", "col": "Precio", "orden": "desc"})
    assert df.iloc[0]["Precio"] == 30.0


def test_ordenar_asc(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {"op": "ordenar", "col": "Cantidad", "orden": "asc"})
    assert df.iloc[0]["Cantidad"] == 2


def test_ordenar_columna_inexistente(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {"op": "ordenar", "col": "NoExiste"})


# ── eliminar_duplicados ───────────────────────────────────────────────────────

def test_eliminar_duplicados_global(df_ventas):
    df_dup = pd.concat([df_ventas, df_ventas.iloc[[0]]], ignore_index=True)
    df, desc, _ = aplicar_edicion(df_dup, {"op": "eliminar_duplicados"})
    assert len(df) == len(df_ventas)
    assert "1" in desc   # "1 fila(s) duplicada(s)"


def test_eliminar_duplicados_por_columna(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "eliminar_duplicados", "columnas": ["Producto"],
    })
    assert len(df) == 3   # A, B, C


# ── filtrar_exportar ──────────────────────────────────────────────────────────

def test_filtrar_exportar(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "filtrar_exportar",
        "filtros": [{"col": "Región", "op": "==", "val": "Norte"}],
    })
    assert len(df) == 3
    assert all(df["Región"] == "Norte")


# ── rellenar_nulos ────────────────────────────────────────────────────────────

def test_rellenar_nulos_media(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "rellenar_nulos", "col": "Descuento", "metodo": "media",
    })
    assert df["Descuento"].isnull().sum() == 0


def test_rellenar_nulos_cero(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "rellenar_nulos", "col": "Descuento", "metodo": "cero",
    })
    assert (df["Descuento"] == 0).sum() == 2   # los 2 nulos se rellenan con 0


def test_rellenar_nulos_valor_fijo(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "rellenar_nulos", "col": "Descuento", "metodo": "valor", "valor": 0.5,
    })
    assert all(df["Descuento"].notna())
    assert (df["Descuento"] == 0.5).sum() == 2


# ── renombrar_columna ─────────────────────────────────────────────────────────

def test_renombrar_columna(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "renombrar_columna",
        "columnas": {"Precio": "PrecioUnitario", "Cantidad": "Uds"},
    })
    assert "PrecioUnitario" in df.columns
    assert "Uds" in df.columns
    assert "Precio" not in df.columns


def test_renombrar_columna_inexistente(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {
            "op": "renombrar_columna", "columnas": {"NoExiste": "X"},
        })


# ── eliminar_columna ──────────────────────────────────────────────────────────

def test_eliminar_columna(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "eliminar_columna", "columnas": ["Descuento"],
    })
    assert "Descuento" not in df.columns
    assert len(df.columns) == 4


def test_eliminar_varias_columnas(df_ventas):
    df, desc, _ = aplicar_edicion(df_ventas, {
        "op": "eliminar_columna", "columnas": ["Descuento", "Región"],
    })
    assert "Descuento" not in df.columns
    assert "Región" not in df.columns


# ── formato_condicional ───────────────────────────────────────────────────────

def test_formato_condicional_no_modifica_datos(df_ventas):
    df_orig = df_ventas.copy()
    df, desc, extras = aplicar_edicion(df_ventas, {
        "op": "formato_condicional",
        "col": "Precio", "condicion": "<", "valor": 15, "color": "rojo",
    })
    assert df.equals(df_orig)   # df no cambia
    assert extras is not None   # instrucciones de formato devueltas
    assert "formato" in desc.lower()


# ── operación desconocida ─────────────────────────────────────────────────────

def test_operacion_desconocida(df_ventas):
    with pytest.raises(EditorError):
        aplicar_edicion(df_ventas, {"op": "inventar"})


# ── exportar_xlsx ─────────────────────────────────────────────────────────────

def test_exportar_genera_xlsx_valido(df_ventas):
    buf, nombre = exportar_xlsx(df_ventas, "test.xlsx", "Prueba de exportación")
    assert nombre == "test_modificado.xlsx"
    assert isinstance(buf, io.BytesIO)
    wb = openpyxl.load_workbook(buf)
    assert "Datos" in wb.sheetnames
    assert "Modificación" in wb.sheetnames


def test_exportar_tiene_cabeceras(df_ventas):
    buf, _ = exportar_xlsx(df_ventas, "test.xlsx")
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb["Datos"]
    cabeceras = [ws.cell(row=1, column=i).value for i in range(1, len(df_ventas.columns) + 1)]
    assert cabeceras == list(df_ventas.columns)


def test_exportar_con_formato_condicional(df_ventas):
    fmt = {"col": "Precio", "condicion": "<", "valor": 15, "color": "rojo"}
    buf, _ = exportar_xlsx(df_ventas, "test.xlsx", formato_condicional=fmt)
    assert buf.getbuffer().nbytes > 0

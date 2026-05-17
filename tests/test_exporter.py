import io
import pytest
import openpyxl
from excel.exporter import crear_ejemplo, crear_plantilla


# ── crear_ejemplo ─────────────────────────────────────────────────────────────

@pytest.mark.parametrize("funcion", ["BUSCARV", "BUSCARX", "SUMAR.SI", "CONTAR.SI", "SI"])
def test_crear_ejemplo_funciones_conocidas(funcion):
    buf, nombre = crear_ejemplo(funcion)
    assert isinstance(buf, io.BytesIO)
    assert buf.getbuffer().nbytes > 0
    assert funcion in nombre.upper()


def test_crear_ejemplo_funcion_desconocida():
    buf, nombre = crear_ejemplo("MIFUNCION")
    assert buf.getbuffer().nbytes > 0
    wb = openpyxl.load_workbook(buf)
    assert len(wb.sheetnames) >= 1


def test_crear_ejemplo_genera_xlsx_valido():
    buf, _ = crear_ejemplo("BUSCARV")
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # Debe tener al menos cabeceras y datos
    assert ws.max_row >= 2
    assert ws.max_column >= 2


# ── crear_plantilla ──────────────────────────────────────────────────────────

@pytest.mark.parametrize("nombre", ["presupuesto", "gastos", "kpis", "inventario"])
def test_crear_plantilla_genera_archivo(nombre):
    buf, fname = crear_plantilla(nombre)
    assert isinstance(buf, io.BytesIO)
    assert buf.getbuffer().nbytes > 0
    assert nombre in fname.lower()


def test_plantilla_presupuesto_tiene_hoja():
    buf, _ = crear_plantilla("presupuesto")
    wb = openpyxl.load_workbook(buf)
    assert "Presupuesto" in wb.sheetnames


def test_plantilla_gastos_tiene_dos_hojas():
    buf, _ = crear_plantilla("gastos")
    wb = openpyxl.load_workbook(buf)
    assert len(wb.sheetnames) == 2


def test_plantilla_inventario_tiene_formulas():
    buf, _ = crear_plantilla("inventario")
    # Cargar sin data_only para ver las fórmulas
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # La columna E debe tener fórmulas (=C*D)
    formulas = [ws.cell(row=r, column=5).value for r in range(2, 7)]
    assert any(str(f).startswith("=") for f in formulas if f)

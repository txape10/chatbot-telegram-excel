import pytest
import openpyxl
import pandas as pd
from excel.reader import leer_excel, leer_excel_hojas, leer_csv


def _excel_simple(tmp_path, datos, titulo="Hoja1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    for fila in datos:
        ws.append(fila)
    ruta = tmp_path / "test.xlsx"
    wb.save(ruta)
    return str(ruta)


# ── leer_excel ───────────────────────────────────────────────────────────────

def test_leer_excel_columnas_y_filas(tmp_path):
    ruta = _excel_simple(tmp_path, [["Nombre", "Edad"], ["Ana", 25], ["Luis", 30]])
    df = leer_excel(ruta)
    assert list(df.columns) == ["Nombre", "Edad"]
    assert len(df) == 2


def test_leer_excel_tipos(tmp_path):
    ruta = _excel_simple(tmp_path, [["Producto", "Precio"], ["Teclado", 29.99]])
    df = leer_excel(ruta)
    assert df["Precio"].dtype in (float, "float64")


def test_leer_excel_con_nulos(tmp_path):
    ruta = _excel_simple(tmp_path, [["A", "B"], [1, None], [2, 3]])
    df = leer_excel(ruta)
    assert df["B"].isnull().sum() == 1


# ── leer_excel_hojas ─────────────────────────────────────────────────────────

def test_leer_excel_hojas_multiples(tmp_path):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Ventas"
    ws1.append(["Producto", "Importe"])
    ws1.append(["A", 100])
    ws2 = wb.create_sheet("Gastos")
    ws2.append(["Concepto", "Importe"])
    ws2.append(["Alquiler", 800])
    ruta = tmp_path / "multi.xlsx"
    wb.save(ruta)

    sheets = leer_excel_hojas(str(ruta))
    assert set(sheets.keys()) == {"Ventas", "Gastos"}
    assert len(sheets["Ventas"]) == 1
    assert len(sheets["Gastos"]) == 1


def test_leer_excel_hojas_una_hoja(tmp_path):
    ruta = _excel_simple(tmp_path, [["X"], [1]])
    sheets = leer_excel_hojas(ruta)
    assert len(sheets) == 1


# ── leer_csv ─────────────────────────────────────────────────────────────────

def test_leer_csv_separador_coma(tmp_path):
    ruta = tmp_path / "datos.csv"
    ruta.write_text("Nombre,Edad\nAna,25\nLuis,30", encoding="utf-8")
    df = leer_csv(str(ruta))
    assert len(df) == 2
    assert "Nombre" in df.columns


def test_leer_csv_separador_punto_coma(tmp_path):
    ruta = tmp_path / "datos.csv"
    ruta.write_text("Nombre;Edad\nAna;25\nLuis;30", encoding="utf-8")
    df = leer_csv(str(ruta))
    assert len(df) == 2
    assert "Nombre" in df.columns


def test_leer_csv_devuelve_dataframe(tmp_path):
    ruta = tmp_path / "datos.csv"
    ruta.write_text("A,B\n1,2\n3,4", encoding="utf-8")
    df = leer_csv(str(ruta))
    assert isinstance(df, pd.DataFrame)

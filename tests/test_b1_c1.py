"""Tests para Sprint B1/B2 (crear_desde_descripcion) y C1/C2 (análisis estadístico)."""
import io
import pytest
import pandas as pd
import openpyxl

from excel.exporter import crear_desde_descripcion
from excel.analyzer import analisis_estadistico_completo, analisis_correlaciones


# ── Fixtures ─────────────────────────────────────────────────────────────────

@pytest.fixture
def df_ventas():
    return pd.DataFrame({
        "Producto": ["Teclado", "Ratón", "Monitor", "Auriculares"],
        "Precio":   [29.99, 19.99, 199.99, 49.99],
        "Cantidad": [10, 25, 5, 12],
        "Ventas":   [299.9, 499.75, 999.95, 599.88],
    })


@pytest.fixture
def df_sin_numericas():
    return pd.DataFrame({
        "Nombre":  ["Ana", "Luis", "Eva"],
        "Ciudad":  ["Madrid", "Barcelona", "Sevilla"],
    })


# ── Tests crear_desde_descripcion (B1/B2) ────────────────────────────────────

def test_crear_desde_descripcion_estructura_basica():
    estructura = {
        "titulo": "Presupuesto",
        "columnas": ["Concepto", "Importe"],
        "datos": [["Alquiler", 800], ["Luz", 60], ["Internet", 30]],
        "agregar_totales": False,
    }
    buf, nombre = crear_desde_descripcion(estructura)
    assert isinstance(buf, io.BytesIO)
    assert nombre == "presupuesto.xlsx"

    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    assert ws.title == "Presupuesto"
    assert ws.cell(1, 1).value == "Concepto"
    assert ws.cell(1, 2).value == "Importe"
    assert ws.cell(2, 1).value == "Alquiler"
    assert ws.cell(4, 2).value == 30


def test_crear_desde_descripcion_con_totales():
    estructura = {
        "titulo": "Ventas",
        "columnas": ["Mes", "Total"],
        "datos": [["Enero", 1000], ["Febrero", 1500], ["Marzo", 1200]],
        "agregar_totales": True,
    }
    buf, _ = crear_desde_descripcion(estructura)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active

    # Fila de totales en fila 5 (1 cabecera + 3 datos + 1 total)
    assert ws.cell(5, 1).value == "TOTAL"
    # La columna Total debe tener una fórmula SUMA
    formula_cell = ws.cell(5, 2).value
    assert formula_cell is not None
    assert "SUMA" in str(formula_cell).upper() or "SUM" in str(formula_cell).upper()


def test_crear_desde_descripcion_sin_datos():
    estructura = {
        "titulo": "Plantilla",
        "columnas": ["Fecha", "Descripción", "Importe"],
        "datos": [],
        "agregar_totales": False,
    }
    buf, nombre = crear_desde_descripcion(estructura)
    wb = openpyxl.load_workbook(buf)
    ws = wb.active
    # Solo la fila de cabecera
    assert ws.cell(1, 1).value == "Fecha"
    assert ws.cell(2, 1).value is None


def test_crear_desde_descripcion_sin_columnas_lanza_error():
    estructura = {"titulo": "Vacío", "columnas": [], "datos": [], "agregar_totales": False}
    with pytest.raises(ValueError):
        crear_desde_descripcion(estructura)


def test_crear_desde_descripcion_titulo_largo_truncado():
    titulo_largo = "A" * 40
    estructura = {
        "titulo": titulo_largo,
        "columnas": ["Col"],
        "datos": [],
        "agregar_totales": False,
    }
    buf, _ = crear_desde_descripcion(estructura)
    wb = openpyxl.load_workbook(buf)
    # Excel limita nombres de hoja a 31 caracteres
    assert len(wb.active.title) <= 31


def test_crear_desde_descripcion_nombre_archivo_normalizado():
    estructura = {
        "titulo": "Mi Presupuesto 2025",
        "columnas": ["Item", "Coste"],
        "datos": [],
        "agregar_totales": False,
    }
    _, nombre = crear_desde_descripcion(estructura)
    assert " " not in nombre
    assert nombre.endswith(".xlsx")


# ── Tests analisis_estadistico_completo (C1) ─────────────────────────────────

def test_estadistico_incluye_todas_columnas_numericas(df_ventas):
    resultado = analisis_estadistico_completo(df_ventas)
    assert "Precio" in resultado
    assert "Cantidad" in resultado
    assert "Ventas" in resultado


def test_estadistico_incluye_metricas_clave(df_ventas):
    resultado = analisis_estadistico_completo(df_ventas)
    assert "Media" in resultado or "media" in resultado.lower()
    assert "Mediana" in resultado or "mediana" in resultado.lower()
    assert "Mín" in resultado or "Min" in resultado or "mín" in resultado.lower()
    assert "Máx" in resultado or "Max" in resultado


def test_estadistico_sin_columnas_numericas(df_sin_numericas):
    resultado = analisis_estadistico_completo(df_sin_numericas)
    assert "no tiene columnas numéricas" in resultado.lower() or "⚠️" in resultado


def test_estadistico_detecta_sesgo():
    # Distribución muy sesgada a la derecha
    df = pd.DataFrame({"Valores": [1, 1, 1, 1, 1, 1, 100, 200, 500]})
    resultado = analisis_estadistico_completo(df)
    assert "sesgada" in resultado.lower() or "derecha" in resultado.lower()


def test_estadistico_ignora_nulos():
    df = pd.DataFrame({"A": [1.0, 2.0, None, 4.0, None]})
    resultado = analisis_estadistico_completo(df)
    # No debe lanzar excepción y debe mostrar resultados
    assert "A" in resultado


# ── Tests analisis_correlaciones (C2) ────────────────────────────────────────

def test_correlaciones_texto_incluye_pares(df_ventas):
    texto, buf_img = analisis_correlaciones(df_ventas)
    assert "correlaciones" in texto.lower() or "↔" in texto


def test_correlaciones_detecta_correlacion_alta():
    df = pd.DataFrame({
        "X": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
        "Y": [2, 4, 6, 8, 10, 12, 14, 16, 18, 20],  # correlación perfecta
    })
    texto, _ = analisis_correlaciones(df)
    assert "1.00" in texto or "muy fuerte" in texto.lower()


def test_correlaciones_genera_imagen(df_ventas):
    _, buf_img = analisis_correlaciones(df_ventas)
    assert buf_img is not None
    assert isinstance(buf_img, io.BytesIO)
    # PNG comienza con la firma \x89PNG
    data = buf_img.read(4)
    assert data[:4] == b'\x89PNG'


def test_correlaciones_insuficientes_columnas(df_sin_numericas):
    texto, buf_img = analisis_correlaciones(df_sin_numericas)
    assert "al menos 2 columnas" in texto.lower() or "⚠️" in texto
    assert buf_img is None


def test_correlaciones_correlacion_negativa():
    df = pd.DataFrame({
        "Precio": [10, 20, 30, 40, 50],
        "Demanda": [500, 400, 300, 200, 100],  # correlación negativa perfecta
    })
    texto, _ = analisis_correlaciones(df)
    assert "negativa" in texto.lower() or "-1" in texto

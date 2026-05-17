import io
import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _estilo_cabecera(celda):
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="2E75B6")
    celda.alignment = Alignment(horizontal="center")


def _estilo_formula(celda):
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="70AD47")
    celda.alignment = Alignment(horizontal="left")


def _ajustar_columnas(hoja):
    for col in hoja.columns:
        ancho = max(len(str(c.value or "")) for c in col) + 4
        hoja.column_dimensions[get_column_letter(col[0].column)].width = min(ancho, 40)


def crear_ejemplo(funcion: str) -> tuple[io.BytesIO, str]:
    """Devuelve (buffer_xlsx, nombre_archivo). Usa ejemplo genérico si la función no está definida."""
    funcion_upper = funcion.upper().replace(" ", ".")
    creador = EJEMPLOS.get(funcion_upper, _ejemplo_generico)
    wb, nombre = creador(funcion_upper)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, nombre


# ── Ejemplos concretos ──────────────────────────────────────────────────────

def _ejemplo_buscarv(funcion: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BUSCARV"

    cabeceras = ["ID", "Producto", "Precio"]
    datos = [(1, "Teclado", 29.99), (2, "Ratón", 19.99), (3, "Monitor", 199.99),
             (4, "Auriculares", 49.99), (5, "Webcam", 39.99)]

    for i, cab in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=i, value=cab)
        _estilo_cabecera(c)
    for fila, dato in enumerate(datos, 2):
        for col, valor in enumerate(dato, 1):
            ws.cell(row=fila, column=col, value=valor)

    ws["E1"] = "ID a buscar"
    _estilo_cabecera(ws["E1"])
    ws["E2"] = 3

    ws["F1"] = "Producto encontrado"
    _estilo_cabecera(ws["F1"])
    c = ws["F2"]
    c.value = '=BUSCARV(E2,$A$2:$C$6,2,FALSO)'
    _estilo_formula(c)

    ws["G1"] = "Precio encontrado"
    _estilo_cabecera(ws["G1"])
    c = ws["G2"]
    c.value = '=BUSCARV(E2,$A$2:$C$6,3,FALSO)'
    _estilo_formula(c)

    _ajustar_columnas(ws)
    return wb, "ejemplo_BUSCARV.xlsx"


def _ejemplo_buscarx(funcion: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BUSCARX"

    cabeceras = ["Código", "Artículo", "Stock", "Precio"]
    datos = [("A001", "Silla", 12, 89.90), ("A002", "Mesa", 5, 149.90),
             ("A003", "Lámpara", 30, 24.50), ("A004", "Estantería", 8, 59.99)]

    for i, cab in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=i, value=cab)
        _estilo_cabecera(c)
    for fila, dato in enumerate(datos, 2):
        for col, valor in enumerate(dato, 1):
            ws.cell(row=fila, column=col, value=valor)

    ws["F1"] = "Código a buscar"
    _estilo_cabecera(ws["F1"])
    ws["F2"] = "A003"

    ws["G1"] = "Resultado BUSCARX"
    _estilo_cabecera(ws["G1"])
    c = ws["G2"]
    c.value = '=BUSCARX(F2,$A$2:$A$5,$C$2:$C$5,"No encontrado")'
    _estilo_formula(c)

    _ajustar_columnas(ws)
    return wb, "ejemplo_BUSCARX.xlsx"


def _ejemplo_sumar_si(funcion: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SUMAR.SI"

    cabeceras = ["Mes", "Vendedor", "Ventas"]
    datos = [("Enero", "Ana", 1200), ("Enero", "Luis", 950), ("Febrero", "Ana", 1400),
             ("Febrero", "Luis", 1100), ("Marzo", "Ana", 1600), ("Marzo", "Luis", 800)]

    for i, cab in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=i, value=cab)
        _estilo_cabecera(c)
    for fila, dato in enumerate(datos, 2):
        for col, valor in enumerate(dato, 1):
            ws.cell(row=fila, column=col, value=valor)

    ws["E1"] = "Vendedor"
    _estilo_cabecera(ws["E1"])
    ws["F1"] = "Total ventas"
    _estilo_cabecera(ws["F1"])

    ws["E2"] = "Ana"
    c = ws["F2"]
    c.value = '=SUMAR.SI($B$2:$B$7,E2,$C$2:$C$7)'
    _estilo_formula(c)

    ws["E3"] = "Luis"
    c = ws["F3"]
    c.value = '=SUMAR.SI($B$2:$B$7,E3,$C$2:$C$7)'
    _estilo_formula(c)

    _ajustar_columnas(ws)
    return wb, "ejemplo_SUMAR.SI.xlsx"


def _ejemplo_contar_si(funcion: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONTAR.SI"

    cabeceras = ["Empleado", "Departamento", "Nota"]
    datos = [("Ana", "Ventas", "Aprobado"), ("Luis", "IT", "Pendiente"),
             ("Sara", "Ventas", "Aprobado"), ("Pedro", "IT", "Aprobado"),
             ("Marta", "Ventas", "Pendiente"), ("Carlos", "IT", "Aprobado")]

    for i, cab in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=i, value=cab)
        _estilo_cabecera(c)
    for fila, dato in enumerate(datos, 2):
        for col, valor in enumerate(dato, 1):
            ws.cell(row=fila, column=col, value=valor)

    ws["E1"] = "Estado"
    _estilo_cabecera(ws["E1"])
    ws["F1"] = "Cantidad"
    _estilo_cabecera(ws["F1"])

    ws["E2"] = "Aprobado"
    c = ws["F2"]
    c.value = '=CONTAR.SI($C$2:$C$7,E2)'
    _estilo_formula(c)

    ws["E3"] = "Pendiente"
    c = ws["F3"]
    c.value = '=CONTAR.SI($C$2:$C$7,E3)'
    _estilo_formula(c)

    _ajustar_columnas(ws)
    return wb, "ejemplo_CONTAR.SI.xlsx"


def _ejemplo_si(funcion: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SI"

    cabeceras = ["Alumno", "Nota", "Resultado", "Calificación"]
    datos = [("Ana", 7.5), ("Luis", 4.2), ("Sara", 9.1), ("Pedro", 5.0), ("Marta", 3.8)]

    for i, cab in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=i, value=cab)
        _estilo_cabecera(c)
    for fila, dato in enumerate(datos, 2):
        ws.cell(row=fila, column=1, value=dato[0])
        ws.cell(row=fila, column=2, value=dato[1])
        c = ws.cell(row=fila, column=3)
        c.value = f'=SI(B{fila}>=5,"Aprobado","Suspenso")'
        _estilo_formula(c)
        c2 = ws.cell(row=fila, column=4)
        c2.value = f'=SI(B{fila}>=9,"Sobresaliente",SI(B{fila}>=7,"Notable",SI(B{fila}>=5,"Aprobado","Suspenso")))'
        _estilo_formula(c2)

    _ajustar_columnas(ws)
    return wb, "ejemplo_SI.xlsx"


def _ejemplo_generico(funcion: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = funcion[:30]

    cabeceras = ["ID", "Categoría", "Valor A", "Valor B", "Resultado"]
    datos = [(1, "Norte", 100, 200), (2, "Sur", 150, 80),
             (3, "Este", 90, 310), (4, "Oeste", 200, 120), (5, "Norte", 75, 95)]

    for i, cab in enumerate(cabeceras, 1):
        c = ws.cell(row=1, column=i, value=cab)
        _estilo_cabecera(c)
    for fila, dato in enumerate(datos, 2):
        for col, valor in enumerate(dato, 1):
            ws.cell(row=fila, column=col, value=valor)
        c = ws.cell(row=fila, column=5)
        c.value = f"Aplica aquí {funcion}"
        _estilo_formula(c)

    ws["G1"] = f"Función: {funcion}"
    ws["G2"] = "Modifica la columna Resultado con tu fórmula"

    _ajustar_columnas(ws)
    return wb, f"ejemplo_{funcion}.xlsx"


# Registro de ejemplos (debe estar después de las funciones)
EJEMPLOS = {
    "BUSCARV": _ejemplo_buscarv,
    "BUSCARX": _ejemplo_buscarx,
    "SUMAR.SI": _ejemplo_sumar_si,
    "CONTAR.SI": _ejemplo_contar_si,
    "SI": _ejemplo_si,
}


# ── Crear desde descripción (B1/B2) ─────────────────────────────────────────

def crear_desde_descripcion(estructura: dict) -> tuple[io.BytesIO, str]:
    """Genera un .xlsx a partir de la estructura JSON extraída por el LLM.

    estructura esperada:
      titulo          → nombre de la hoja
      columnas        → lista de nombres de columna
      datos           → lista de listas con los valores (puede ser vacía)
      agregar_totales → bool, si añadir fila de SUMA al final
    """
    titulo   = estructura.get("titulo", "Datos")[:31]   # Excel limita a 31 chars
    columnas = estructura.get("columnas", [])
    datos    = estructura.get("datos", [])
    totales  = estructura.get("agregar_totales", False)

    if not columnas:
        raise ValueError("La estructura no contiene columnas.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo

    # Cabeceras
    for ci, col in enumerate(columnas, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        _estilo_cabecera(c)

    # Datos
    for ri, fila in enumerate(datos, 2):
        for ci, valor in enumerate(fila, 1):
            ws.cell(row=ri, column=ci, value=valor)

    # Fila de totales (SUMA en columnas numéricas)
    if totales and datos:
        fila_total = len(datos) + 2
        ws.cell(row=fila_total, column=1, value="TOTAL")
        ws.cell(row=fila_total, column=1).font = Font(bold=True)

        for ci in range(1, len(columnas) + 1):
            # Detectar si la columna tiene valores numéricos
            valores_col = [
                ws.cell(row=r, column=ci).value
                for r in range(2, fila_total)
            ]
            if any(isinstance(v, (int, float)) for v in valores_col if v is not None):
                col_letra = get_column_letter(ci)
                c = ws.cell(row=fila_total, column=ci)
                c.value = f"=SUMA({col_letra}2:{col_letra}{fila_total - 1})"
                _estilo_formula(c)

    _ajustar_columnas(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"{titulo.lower().replace(' ', '_')}.xlsx"
    return buf, nombre


# ── Tabla dinámica ───────────────────────────────────────────────────────────

def crear_tabla_dinamica(df: pd.DataFrame | None = None) -> tuple[io.BytesIO, str]:
    """Genera un .xlsx con dos hojas: datos fuente + resumen tipo tabla dinámica.

    Si se pasa un DataFrame del usuario se usan sus datos;
    si no, se genera un ejemplo con datos de ventas ficticios.
    """
    usar_datos_usuario = df is not None and not df.empty

    if not usar_datos_usuario:
        df = pd.DataFrame({
            "Mes":        ["Enero",   "Enero",   "Enero",    "Febrero", "Febrero",
                           "Febrero", "Marzo",   "Marzo",    "Marzo",   "Abril"],
            "Categoría":  ["Ventas",  "Ventas",  "Servicios","Ventas",  "Servicios",
                           "Ventas",  "Servicios","Ventas",  "Ventas",  "Servicios"],
            "Región":     ["Norte",   "Sur",     "Norte",    "Sur",     "Norte",
                           "Norte",   "Sur",     "Norte",    "Sur",     "Norte"],
            "Importe":    [1200, 950, 400, 1400, 600, 1100, 500, 1300, 800, 450],
            "Unidades":   [12,   9,   4,   14,   6,   11,   5,   13,   8,   4],
        })
        nombre_archivo = "tabla_dinamica_ejemplo.xlsx"
    else:
        nombre_archivo = "tabla_dinamica_tus_datos.xlsx"

    # ── Detectar columnas categóricas y numéricas ─────────────────────────────
    # Categóricas: texto o cualquier columna con pocos valores únicos relativos al total
    umbral_cat = max(10, len(df) // 5)  # flexible según tamaño del df
    cols_cat = [
        c for c in df.columns
        if not pd.api.types.is_numeric_dtype(df[c])
        or (df[c].nunique() <= umbral_cat and df[c].nunique() < len(df))
    ]
    cols_num = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    # Fallback: si no hay categorías claras, usar la columna con menos valores únicos
    if not cols_cat and len(df.columns) > 0:
        cols_cat = [min(df.columns, key=lambda c: df[c].nunique())]

    # Excluir de cols_num las que ya usamos como categóricas
    cols_num = [c for c in cols_num if c not in cols_cat]
    # Si todas las numéricas quedaron como categóricas, usarlas también como numéricas
    if not cols_num:
        cols_num = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]

    wb = openpyxl.Workbook()

    # ── Hoja 1: Datos fuente ──────────────────────────────────────────────────
    ws_datos = wb.active
    ws_datos.title = "Datos"

    for i, col in enumerate(df.columns, 1):
        _estilo_cabecera(ws_datos.cell(row=1, column=i, value=col))
    for fila_idx, fila in enumerate(df.itertuples(index=False), 2):
        for col_idx, valor in enumerate(fila, 1):
            ws_datos.cell(row=fila_idx, column=col_idx, value=valor)
    _ajustar_columnas(ws_datos)

    # ── Hoja 2: Resúmenes ─────────────────────────────────────────────────────
    ws_td = wb.create_sheet("Tabla Dinámica")

    fila_actual = 1
    borde = Border(
        bottom=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
    )

    # Resumen 1: por primera columna categórica × columnas numéricas (máx. 3)
    if cols_cat and cols_num:
        col_fila = cols_cat[0]
        cols_agg = cols_num[:3]

        try:
            pivot = (
                df.groupby(col_fila)[cols_agg]
                .agg(["sum", "mean"])
                .round(2)
            )
            pivot.columns = [f"{agg.capitalize()} {col}" for col, agg in pivot.columns]
            pivot = pivot.reset_index().sort_values(pivot.columns[1], ascending=False)

            titulo_txt = f"Resumen por {col_fila}"
            titulo = ws_td.cell(row=fila_actual, column=1, value=titulo_txt)
            titulo.font = Font(bold=True, color="FFFFFF", size=12)
            titulo.fill = PatternFill("solid", fgColor="1F3864")
            ws_td.merge_cells(start_row=fila_actual, start_column=1,
                              end_row=fila_actual, end_column=len(pivot.columns))
            fila_actual += 1

            for ci, cab in enumerate(pivot.columns, 1):
                _estilo_cabecera(ws_td.cell(row=fila_actual, column=ci, value=str(cab)))
            fila_actual += 1

            for ri, row in enumerate(pivot.itertuples(index=False)):
                for ci, valor in enumerate(row, 1):
                    c = ws_td.cell(row=fila_actual, column=ci, value=valor)
                    if ri % 2 == 0:
                        c.fill = PatternFill("solid", fgColor="EBF3FB")
                    c.border = borde
                    if isinstance(valor, float):
                        c.number_format = "#,##0.00"
                fila_actual += 1

            # Fila de totales para columnas numéricas
            c_label = ws_td.cell(row=fila_actual, column=1, value="TOTAL")
            _estilo_total(c_label)
            for ci, col_name in enumerate(pivot.columns[1:], 2):
                orig_col = cols_agg[(ci - 2) % len(cols_agg)]
                c = ws_td.cell(row=fila_actual, column=ci)
                if "sum" in col_name.lower():
                    c.value = round(float(df[orig_col].sum()), 2)
                else:
                    c.value = round(float(df[orig_col].mean()), 2)
                _estilo_total(c)
                c.number_format = "#,##0.00"
            fila_actual += 2

        except Exception:
            pass  # si el groupby falla por cualquier razón, seguimos con el siguiente bloque

    # Resumen 2: cruce de dos categorías (si hay suficientes)
    if len(cols_cat) >= 2 and cols_num:
        col_fila   = cols_cat[0]
        col_columna = cols_cat[1]
        col_val    = cols_num[0]

        try:
            crosstab = df.pivot_table(
                values=col_val,
                index=col_fila,
                columns=col_columna,
                aggfunc="sum",
                fill_value=0,
            ).reset_index()

            # Omitir si el cruce genera demasiadas columnas (ilegible)
            if len(crosstab.columns) <= 12:
                titulo = ws_td.cell(row=fila_actual, column=1,
                                    value=f"Cruce: {col_fila} × {col_columna} (suma de {col_val})")
                titulo.font = Font(bold=True, color="FFFFFF", size=12)
                titulo.fill = PatternFill("solid", fgColor="385723")
                ws_td.merge_cells(start_row=fila_actual, start_column=1,
                                  end_row=fila_actual, end_column=len(crosstab.columns))
                fila_actual += 1

                for ci, col_name in enumerate(crosstab.columns, 1):
                    _estilo_cabecera(ws_td.cell(row=fila_actual, column=ci, value=str(col_name)))
                fila_actual += 1

                for ri, row in enumerate(crosstab.itertuples(index=False)):
                    for ci, valor in enumerate(row, 1):
                        c = ws_td.cell(row=fila_actual, column=ci, value=valor)
                        if ri % 2 == 0:
                            c.fill = PatternFill("solid", fgColor="E2EFDA")
                        c.border = borde
                        if isinstance(valor, (int, float)) and ci > 1:
                            c.number_format = "#,##0.00"
                    fila_actual += 1
                fila_actual += 1

        except Exception:
            pass

    # Resumen 3 (fallback): estadísticas descriptivas si no hubo agrupaciones
    if fila_actual == 1 and cols_num:
        titulo = ws_td.cell(row=fila_actual, column=1, value="Estadísticas descriptivas")
        titulo.font = Font(bold=True, color="FFFFFF", size=12)
        titulo.fill = PatternFill("solid", fgColor="1F3864")
        ws_td.merge_cells(start_row=fila_actual, start_column=1,
                          end_row=fila_actual, end_column=5)
        fila_actual += 1

        for ci, cab in enumerate(["Columna", "Suma", "Promedio", "Mín", "Máx"], 1):
            _estilo_cabecera(ws_td.cell(row=fila_actual, column=ci, value=cab))
        fila_actual += 1

        for ri, col in enumerate(cols_num):
            serie = df[col].dropna()
            vals = [col, round(float(serie.sum()), 2), round(float(serie.mean()), 2),
                    round(float(serie.min()), 2), round(float(serie.max()), 2)]
            for ci, v in enumerate(vals, 1):
                c = ws_td.cell(row=fila_actual, column=ci, value=v)
                if ri % 2 == 0:
                    c.fill = PatternFill("solid", fgColor="EBF3FB")
                c.border = borde
                if isinstance(v, float):
                    c.number_format = "#,##0.00"
            fila_actual += 1
        fila_actual += 1

    # Nota explicativa
    nota = ws_td.cell(row=fila_actual, column=1,
                      value="💡 En Excel puedes crear tablas dinámicas interactivas desde: "
                            "Insertar → Tabla dinámica. Este archivo muestra el resultado equivalente.")
    nota.font = Font(italic=True, color="595959")

    _ajustar_columnas(ws_td)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, nombre_archivo


# ── Plantillas de uso ────────────────────────────────────────────────────────

def crear_plantilla(nombre: str) -> tuple[io.BytesIO, str]:
    """Genera una plantilla lista para usar. nombre: presupuesto|gastos|kpis|inventario."""
    creadores = {
        "presupuesto": _plantilla_presupuesto,
        "gastos":      _plantilla_gastos,
        "kpis":        _plantilla_kpis,
        "inventario":  _plantilla_inventario,
    }
    creador = creadores.get(nombre, _plantilla_presupuesto)
    wb, nombre_archivo = creador()
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, nombre_archivo


def _estilo_total(celda):
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="ED7D31")
    celda.alignment = Alignment(horizontal="right")


def _plantilla_presupuesto():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto"

    cabeceras = ["Categoría", "Presupuesto", "Real", "Diferencia", "% Cumplido"]
    for i, cab in enumerate(cabeceras, 1):
        _estilo_cabecera(ws.cell(row=1, column=i, value=cab))

    categorias = ["Vivienda", "Alimentación", "Transporte", "Salud", "Ocio",
                  "Ropa", "Educación", "Ahorro", "Otros"]
    for fila, cat in enumerate(categorias, 2):
        ws.cell(row=fila, column=1, value=cat)
        ws.cell(row=fila, column=2, value=0)   # Presupuesto — el usuario lo rellena
        ws.cell(row=fila, column=3, value=0)   # Real — el usuario lo rellena
        c_dif = ws.cell(row=fila, column=4)
        c_dif.value = f"=B{fila}-C{fila}"
        _estilo_formula(c_dif)
        c_pct = ws.cell(row=fila, column=5)
        c_pct.value = f'=SI(B{fila}=0,"—",C{fila}/B{fila})'
        c_pct.number_format = "0.0%"
        _estilo_formula(c_pct)

    fila_total = len(categorias) + 2
    ws.cell(row=fila_total, column=1, value="TOTAL")
    letras = {2: "B", 3: "C", 4: "D"}
    for col in range(2, 5):
        c = ws.cell(row=fila_total, column=col)
        c.value = f"=SUMA({letras[col]}2:{letras[col]}{fila_total - 1})"
        _estilo_total(c)

    _ajustar_columnas(ws)
    return wb, "plantilla_presupuesto.xlsx"


def _plantilla_gastos():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gastos"

    cabeceras = ["Fecha", "Descripción", "Categoría", "Importe", "Acumulado"]
    for i, cab in enumerate(cabeceras, 1):
        _estilo_cabecera(ws.cell(row=1, column=i, value=cab))

    import datetime
    hoy = datetime.date.today()
    ejemplos = [
        (hoy, "Supermercado", "Alimentación", 65.30),
        (hoy, "Gasolina", "Transporte", 48.00),
        (hoy, "Netflix", "Ocio", 15.99),
    ]
    for fila, (fecha, desc, cat, imp) in enumerate(ejemplos, 2):
        ws.cell(row=fila, column=1, value=fecha).number_format = "dd/mm/yyyy"
        ws.cell(row=fila, column=2, value=desc)
        ws.cell(row=fila, column=3, value=cat)
        ws.cell(row=fila, column=4, value=imp)
        c = ws.cell(row=fila, column=5)
        c.value = f"=SUMA($D$2:D{fila})"
        _estilo_formula(c)

    # Hoja de resumen por categoría
    ws2 = wb.create_sheet("Resumen")
    cabeceras2 = ["Categoría", "Total gastado"]
    for i, cab in enumerate(cabeceras2, 1):
        _estilo_cabecera(ws2.cell(row=1, column=i, value=cab))
    cats = ["Alimentación", "Transporte", "Ocio", "Salud", "Vivienda", "Otros"]
    for fila, cat in enumerate(cats, 2):
        ws2.cell(row=fila, column=1, value=cat)
        c = ws2.cell(row=fila, column=2)
        c.value = f'=SUMAR.SI(Gastos!$C:$C,A{fila},Gastos!$D:$D)'
        _estilo_formula(c)

    _ajustar_columnas(ws)
    _ajustar_columnas(ws2)
    return wb, "plantilla_gastos.xlsx"


def _plantilla_kpis():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPIs"

    cabeceras = ["KPI", "Objetivo", "Real", "Cumplimiento", "Estado"]
    for i, cab in enumerate(cabeceras, 1):
        _estilo_cabecera(ws.cell(row=1, column=i, value=cab))

    kpis = [
        ("Ventas mensuales (€)", 50000, 0),
        ("Nuevos clientes", 20, 0),
        ("Tasa de conversión (%)", 15, 0),
        ("Satisfacción cliente (1-10)", 8, 0),
        ("Tickets resueltos", 100, 0),
        ("Coste por adquisición (€)", 30, 0),
    ]
    for fila, (kpi, obj, real) in enumerate(kpis, 2):
        ws.cell(row=fila, column=1, value=kpi)
        ws.cell(row=fila, column=2, value=obj)
        ws.cell(row=fila, column=3, value=real)
        c_pct = ws.cell(row=fila, column=4)
        c_pct.value = f'=SI(B{fila}=0,"—",C{fila}/B{fila})'
        c_pct.number_format = "0.0%"
        _estilo_formula(c_pct)
        c_est = ws.cell(row=fila, column=5)
        c_est.value = f'=SI(C{fila}=0,"Sin datos",SI(C{fila}>=B{fila},"✅ OK","⚠️ Por debajo"))'
        _estilo_formula(c_est)

    _ajustar_columnas(ws)
    return wb, "plantilla_kpis.xlsx"


def _plantilla_inventario():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    cabeceras = ["Código", "Producto", "Cantidad", "Precio unit.", "Valor total", "Stock mín.", "Alerta"]
    for i, cab in enumerate(cabeceras, 1):
        _estilo_cabecera(ws.cell(row=1, column=i, value=cab))

    productos = [
        ("P001", "Teclado mecánico", 15, 45.99, 5),
        ("P002", "Ratón inalámbrico", 8, 22.50, 10),
        ("P003", "Monitor 24\"", 3, 189.00, 2),
        ("P004", "Auriculares USB", 20, 35.00, 5),
        ("P005", "Webcam HD", 6, 49.99, 4),
    ]
    for fila, (cod, prod, qty, precio, stock_min) in enumerate(productos, 2):
        ws.cell(row=fila, column=1, value=cod)
        ws.cell(row=fila, column=2, value=prod)
        ws.cell(row=fila, column=3, value=qty)
        ws.cell(row=fila, column=4, value=precio)
        c_val = ws.cell(row=fila, column=5)
        c_val.value = f"=C{fila}*D{fila}"
        _estilo_formula(c_val)
        ws.cell(row=fila, column=6, value=stock_min)
        c_alert = ws.cell(row=fila, column=7)
        c_alert.value = f'=SI(C{fila}<=F{fila},"🔴 Reponer","✅ OK")'
        _estilo_formula(c_alert)

    fila_total = len(productos) + 2
    ws.cell(row=fila_total, column=2, value="VALOR TOTAL STOCK")
    c_tot = ws.cell(row=fila_total, column=5)
    c_tot.value = f"=SUMA(E2:E{fila_total - 1})"
    _estilo_total(c_tot)

    _ajustar_columnas(ws)
    return wb, "plantilla_inventario.xlsx"

"""Editor de DataFrames y exportador a .xlsx.

Aplica operaciones de modificación sobre el archivo del usuario y devuelve
el resultado como un nuevo .xlsx listo para descargar.

Operaciones: añadir_columna, ordenar, eliminar_duplicados, filtrar_exportar,
             rellenar_nulos, renombrar_columna, eliminar_columna, formato_condicional,
             normalizar_texto, estandarizar_fechas, despivotear, pivotear.
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_COLORES_HEX = {
    "rojo":     "FF6B6B",
    "verde":    "70AD47",
    "amarillo": "FFD966",
    "naranja":  "F4B942",
    "azul":     "5B9BD5",
}

_OPERADORES_ARITMETICOS = {"+", "-", "*", "/"}

# Funciones numéricas que el LLM puede generar como operador de añadir_columna.
# Cada función recibe (serie_pandas, valor_fijo_o_None) y devuelve una serie.
_FUNCIONES_COLUMNA: dict[str, object] = {
    "redondear": lambda s, v: s.round(int(v) if v is not None else 0),
    "round":     lambda s, v: s.round(int(v) if v is not None else 0),
    "abs":       lambda s, _: s.abs(),
    "absoluto":  lambda s, _: s.abs(),
    "raiz":      lambda s, _: s.pow(0.5),
    "sqrt":      lambda s, _: s.pow(0.5),
    "potencia":  lambda s, v: s.pow(float(v) if v is not None else 2),
    "pow":       lambda s, v: s.pow(float(v) if v is not None else 2),
    "entero":    lambda s, _: s.apply(lambda x: int(x) if pd.notna(x) else x),
    "truncar":   lambda s, _: s.apply(lambda x: int(x) if pd.notna(x) else x),
}

_COMPARADORES = {
    "<":  lambda a, b: a < b,
    ">":  lambda a, b: a > b,
    "<=": lambda a, b: a <= b,
    ">=": lambda a, b: a >= b,
    "==": lambda a, b: a == b,
    "!=": lambda a, b: a != b,
}


class EditorError(Exception):
    """Error controlado al aplicar una operación de edición."""


# ── Dispatcher principal ──────────────────────────────────────────────────────

def aplicar_edicion(df: pd.DataFrame, op: dict,
                    copy_df: bool = True) -> tuple[pd.DataFrame, str, dict | None]:
    """Aplica una operación de edición al DataFrame.

    Devuelve (df_modificado, descripcion, extras).
    extras puede contener instrucciones de formato condicional para exportar_xlsx.

    copy_df=False: el llamador ya hizo la copia defensiva (evita N copias en pipelines).
    """
    tipo = str(op.get("op", "")).lower().strip()
    if copy_df:
        df = df.copy()

    if tipo == "añadir_columna":
        df, desc = _añadir_columna(df, op)
        return df, desc, None

    elif tipo == "ordenar":
        df, desc = _ordenar(df, op)
        return df, desc, None

    elif tipo == "eliminar_duplicados":
        df, desc = _eliminar_duplicados(df, op)
        return df, desc, None

    elif tipo == "filtrar_exportar":
        df, desc = _filtrar_exportar(df, op)
        return df, desc, None

    elif tipo == "rellenar_nulos":
        df, desc = _rellenar_nulos(df, op)
        return df, desc, None

    elif tipo == "renombrar_columna":
        df, desc = _renombrar_columna(df, op)
        return df, desc, None

    elif tipo == "eliminar_columna":
        df, desc = _eliminar_columna(df, op)
        return df, desc, None

    elif tipo == "formato_condicional":
        desc = _describir_formato(op)
        return df, desc, op   # df sin cambios; extras lleva las instrucciones

    elif tipo == "normalizar_texto":
        df, desc = _normalizar_texto(df, op)
        return df, desc, None

    elif tipo == "estandarizar_fechas":
        df, desc = _estandarizar_fechas(df, op)
        return df, desc, None

    elif tipo == "despivotear":
        df, desc = _despivotear(df, op)
        return df, desc, None

    elif tipo == "pivotear":
        df, desc = _pivotear(df, op)
        return df, desc, None

    elif tipo == "buscar_reemplazar":
        df, desc = _buscar_reemplazar(df, op)
        return df, desc, None

    elif tipo == "dividir_columna":
        df, desc = _dividir_columna(df, op)
        return df, desc, None

    elif tipo == "concatenar_columnas":
        df, desc = _concatenar_columnas(df, op)
        return df, desc, None

    elif tipo == "añadir_fila_total":
        df, desc = _añadir_fila_total(df, op)
        return df, desc, None

    elif tipo == "duplicar_filas":
        df, desc = _duplicar_filas(df, op)
        return df, desc, None

    elif tipo == "transponer":
        df, desc = _transponer(df, op)
        return df, desc, None

    else:
        raise EditorError(f"Operación no reconocida: '{tipo}'")


# ── Operaciones ───────────────────────────────────────────────────────────────

def _normalizar_nombre_columna(nombre: str, df: pd.DataFrame) -> str:
    """Ajusta la capitalización del nombre nuevo para que concuerde con el resto de columnas."""
    cols = [str(c) for c in df.columns]
    mayusculas = sum(1 for c in cols if c and c[0].isupper())
    minusculas  = sum(1 for c in cols if c and c[0].islower())
    if mayusculas >= minusculas and nombre:
        return nombre[0].upper() + nombre[1:]
    return nombre


def _añadir_columna(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    nombre = op.get("nombre")
    if not nombre:
        raise EditorError("Falta el nombre de la columna nueva ('nombre').")

    # Ajustar capitalización para que concuerde con las columnas existentes
    nombre = _normalizar_nombre_columna(nombre, df)

    expresion = op.get("expresion")
    if expresion:
        try:
            df[nombre] = df.eval(expresion)
            return df, f"Columna '{nombre}' añadida ({expresion})"
        except Exception as exc:
            raise EditorError(f"Expresión no válida '{expresion}': {exc}") from exc

    col1     = op.get("col1")
    col2     = op.get("col2")
    operador = op.get("operador")
    valor_fijo = op.get("valor_fijo")

    if not col1 or not operador:
        raise EditorError("Faltan 'col1' y 'operador' para añadir_columna.")
    _validar_col(df, col1)

    s1 = pd.to_numeric(df[col1], errors="coerce").fillna(0)
    op_lower = operador.lower()

    # ── Funciones numéricas (redondear, abs, raiz, …) ────────────────────────
    if op_lower in _FUNCIONES_COLUMNA:
        fn = _FUNCIONES_COLUMNA[op_lower]
        df[nombre] = fn(s1, valor_fijo)
        arg = f", {valor_fijo}" if valor_fijo is not None else ""
        return df, f"Columna '{nombre}' añadida ({operador}({col1}{arg}))"

    # ── Operadores aritméticos (+, -, *, /) ──────────────────────────────────
    if operador not in _OPERADORES_ARITMETICOS:
        raise EditorError(
            f"Operador no permitido: '{operador}'. "
            f"Usa aritmético (+, -, *, /) o función ({', '.join(_FUNCIONES_COLUMNA)})"
        )

    if col2:
        _validar_col(df, col2)
        s2 = pd.to_numeric(df[col2], errors="coerce").fillna(0)
    elif valor_fijo is not None:
        s2 = float(valor_fijo)
    else:
        raise EditorError("Falta 'col2' o 'valor_fijo' para la operación.")

    if operador == "+":
        df[nombre] = s1 + s2
    elif operador == "-":
        df[nombre] = s1 - s2
    elif operador == "*":
        df[nombre] = s1 * s2
    elif operador == "/":
        df[nombre] = s1 / s2

    ref = col2 if col2 else str(valor_fijo)
    return df, f"Columna '{nombre}' añadida ({col1} {operador} {ref})"


def _ordenar(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    col = op.get("col")
    _validar_col(df, col)
    asc = str(op.get("orden", "asc")).lower() == "asc"
    # Detectar columnas de fecha (DD/MM/YYYY u otros formatos) para ordenar correctamente
    try:
        temp = pd.to_datetime(df[col], dayfirst=True, errors="coerce")
        if temp.notna().sum() > len(df) * 0.5:
            df = df.assign(_sort_key=temp).sort_values("_sort_key", ascending=asc).drop(columns=["_sort_key"]).reset_index(drop=True)
        else:
            df = df.sort_values(col, ascending=asc).reset_index(drop=True)
    except Exception:
        df = df.sort_values(col, ascending=asc).reset_index(drop=True)
    return df, f"Ordenado por '{col}' {'↑ ascendente' if asc else '↓ descendente'}"


def _eliminar_duplicados(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    columnas = op.get("columnas") or None
    if columnas:
        for c in columnas:
            _validar_col(df, c)
    antes = len(df)
    df = df.drop_duplicates(subset=columnas).reset_index(drop=True)
    eliminadas = antes - len(df)
    return df, f"{eliminadas} fila(s) duplicada(s) eliminada(s)"


def _filtrar_exportar(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    from excel.query_engine import _aplicar_filtros
    filtros = op.get("filtros", [])
    if not filtros:
        raise EditorError("Falta 'filtros' para filtrar_exportar.")
    try:
        df = _aplicar_filtros(df, filtros).reset_index(drop=True)
    except (TypeError, ValueError) as e:
        raise EditorError(f"Error al aplicar filtro: {e}") from e
    return df, f"{len(df)} filas tras aplicar los filtros"


def _rellenar_nulos(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    col    = op.get("col")
    metodo = str(op.get("metodo", "media")).lower()
    cols   = [col] if col else [c for c in df.columns if df[c].isnull().any()]

    if col:
        _validar_col(df, col)

    rellenadas = 0
    for c in cols:
        nulos_antes = int(df[c].isnull().sum())
        if nulos_antes == 0:
            continue
        serie_num = pd.to_numeric(df[c], errors="coerce")
        if metodo == "media":
            df[c] = df[c].fillna(serie_num.mean())
        elif metodo == "mediana":
            df[c] = df[c].fillna(serie_num.median())
        elif metodo == "cero":
            df[c] = df[c].fillna(0)
        elif metodo == "anterior":
            df[c] = df[c].ffill()
        elif metodo == "siguiente":
            df[c] = df[c].bfill()
        elif metodo == "valor":
            df[c] = df[c].fillna(op.get("valor", 0))
        else:
            raise EditorError(f"Método de relleno no reconocido: '{metodo}'")
        rellenadas += nulos_antes

    return df, f"{rellenadas} celda(s) vacía(s) rellenada(s) (método: {metodo})"


def _renombrar_columna(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    mapeo = op.get("columnas", {})
    if not mapeo:
        raise EditorError("Falta 'columnas' {{nombre_actual: nombre_nuevo}}.")
    for c in mapeo:
        _validar_col(df, c)
    df = df.rename(columns=mapeo)
    cambios = ", ".join(f"'{v}' → '{n}'" for v, n in mapeo.items())
    return df, f"Columnas renombradas: {cambios}"


def _eliminar_columna(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    columnas = op.get("columnas", [])
    if not columnas:
        raise EditorError("Falta 'columnas' (lista de nombres a eliminar).")
    for c in columnas:
        _validar_col(df, c)
    df = df.drop(columns=columnas)
    return df, f"Eliminada(s): {', '.join(columnas)}"


def _normalizar_texto(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Limpia espacios y/o unifica capitalización en columnas de texto."""
    col    = op.get("col")
    accion = str(op.get("accion", "todas")).lower().strip()

    if col:
        _validar_col(df, col)
        cols_obj = [col]
    else:
        cols_obj = [c for c in df.columns if df[c].dtype == object]

    if not cols_obj:
        raise EditorError("No hay columnas de texto a las que aplicar la normalización.")

    _ACCIONES = {"strip", "upper", "lower", "title", "todas"}
    if accion not in _ACCIONES:
        raise EditorError(f"Acción no reconocida: '{accion}'. Usa: {', '.join(_ACCIONES)}")

    aplicadas = 0
    for c in cols_obj:
        if df[c].dtype != object:
            continue
        if accion in ("strip", "todas"):
            df[c] = df[c].str.strip()
        if accion in ("upper",):
            df[c] = df[c].str.upper()
        elif accion in ("lower",):
            df[c] = df[c].str.lower()
        elif accion in ("title",):
            df[c] = df[c].str.title()
        # "todas" aplica strip + title como limpieza estándar
        if accion == "todas":
            df[c] = df[c].str.title()
        aplicadas += 1

    sufijo = f"columna '{col}'" if col else f"{aplicadas} columna(s) de texto"
    return df, f"Texto normalizado en {sufijo} (acción: {accion})"


def _estandarizar_fechas(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Convierte columnas de texto con fechas a formato estándar DD/MM/YYYY."""
    col = op.get("col")
    fmt_salida = op.get("formato_salida", "%d/%m/%Y")

    _PALABRAS_FECHA = ("fecha", "date", "dia", "día", "mes", "año", "year", "month",
                       "periodo", "semana", "trimestre")

    if col:
        _validar_col(df, col)
        cols_fecha = [col]
    else:
        cols_fecha = [c for c in df.columns
                      if any(p in str(c).lower() for p in _PALABRAS_FECHA)]

    if not cols_fecha:
        raise EditorError(
            "No encontré columnas de fecha. Especifica la columna con 'col'."
        )

    convertidas = 0
    errores_total = 0
    for c in cols_fecha:
        # format="mixed" permite parsear múltiples formatos de fecha en la misma columna
        try:
            serie_dt = pd.to_datetime(df[c], errors="coerce", dayfirst=True, format="mixed")
        except TypeError:
            serie_dt = pd.to_datetime(df[c], errors="coerce", dayfirst=True)
        validas  = serie_dt.notna()
        errores  = int(validas.sum() == 0)
        if validas.any():
            df.loc[validas, c] = serie_dt[validas].dt.strftime(fmt_salida)
            convertidas += 1
        errores_total += int((~validas & df[c].notna()).sum())

    desc = f"Fechas estandarizadas a {fmt_salida} en {convertidas} columna(s)"
    if errores_total:
        desc += f" ({errores_total} valor(es) no reconocido(s) sin cambios)"
    return df, desc


def _despivotear(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Convierte columnas de valores en filas (pd.melt / unpivot)."""
    cols_valores = op.get("columnas_valores", [])
    if not cols_valores:
        raise EditorError("Falta 'columnas_valores' (lista de columnas a convertir en filas).")
    for c in cols_valores:
        _validar_col(df, c)

    cols_id = op.get("columnas_id") or [c for c in df.columns if c not in cols_valores]
    for c in cols_id:
        _validar_col(df, c)

    col_nombre = op.get("col_nombre", "Variable")
    col_valor  = op.get("col_valor",  "Valor")

    df_melt = df.melt(
        id_vars=cols_id,
        value_vars=cols_valores,
        var_name=col_nombre,
        value_name=col_valor,
    ).reset_index(drop=True)

    return df_melt, (
        f"Despivotado: {len(cols_valores)} columna(s) → filas "
        f"('{col_nombre}' / '{col_valor}'). "
        f"Resultado: {len(df_melt)} filas × {len(df_melt.columns)} columnas"
    )


def _pivotear(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Convierte filas en columnas (pd.pivot_table)."""
    index   = op.get("index")
    columns = op.get("columns")
    values  = op.get("values")

    if not index or not columns or not values:
        raise EditorError("Faltan 'index', 'columns' y/o 'values' para pivotear.")
    for c in (index, columns, values):
        _validar_col(df, c)

    _AGGFUNCS = {
        "suma":    "sum",
        "sum":     "sum",
        "promedio":"mean",
        "mean":    "mean",
        "contar":  "count",
        "count":   "count",
        "max":     "max",
        "min":     "min",
    }
    aggfunc_str = str(op.get("aggfunc", "suma")).lower()
    aggfunc = _AGGFUNCS.get(aggfunc_str, "sum")

    try:
        df_pivot = df.pivot_table(
            index=index, columns=columns, values=values,
            aggfunc=aggfunc, fill_value=0,
        ).reset_index()
        df_pivot.columns = [str(c) for c in df_pivot.columns]
    except Exception as e:
        raise EditorError(f"No se pudo pivotear: {e}") from e

    return df_pivot, (
        f"Pivotado: filas={index}, columnas={columns}, valores={values} ({aggfunc_str}). "
        f"Resultado: {len(df_pivot)} filas × {len(df_pivot.columns)} columnas"
    )


def _describir_formato(op: dict) -> str:
    col      = op.get("col", "?")
    cond     = op.get("condicion", "?")
    valor    = op.get("valor", "?")
    color    = op.get("color", "rojo")
    return f"Formato condicional: '{col}' {cond} {valor} → {color}"


# ── Exportador ────────────────────────────────────────────────────────────────

def exportar_xlsx(df: pd.DataFrame, nombre_base: str,
                  descripcion: str = "",
                  formato_condicional: dict | None = None) -> tuple[io.BytesIO, str]:
    """Convierte el DataFrame modificado a .xlsx con cabeceras formateadas.

    Devuelve (buffer, nombre_archivo).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos"

    # Cabeceras
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        c.font      = Font(bold=True, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center")

    # Datos (con filas alternas ligeramente coloreadas)
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, valor in enumerate(row, 1):
            val = None if (isinstance(valor, float) and pd.isna(valor)) else valor
            c = ws.cell(row=ri, column=ci, value=val)
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="EBF3FB")

    # Formato condicional (coloreado de celdas)
    if formato_condicional:
        _aplicar_formato_condicional(ws, df, formato_condicional)

    # Ajustar anchos de columna
    for col_cells in ws.columns:
        ancho = max((len(str(c.value or "")) for c in col_cells), default=8) + 3
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(ancho, 40)

    # Hoja de información si hay descripción
    if descripcion:
        ws_info = wb.create_sheet("Modificación")
        ws_info["A1"] = "Operación aplicada"
        ws_info["A1"].font = Font(bold=True)
        ws_info["A2"] = descripcion
        ws_info.column_dimensions["A"].width = max(len(descripcion) + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    base = nombre_base.replace(".xlsx", "").replace(".xls", "").replace(".csv", "")
    return buf, f"{base}_modificado.xlsx"


def _aplicar_formato_condicional(ws, df: pd.DataFrame, op: dict) -> None:
    col      = op.get("col")
    condicion = op.get("condicion", "<")
    valor_ref = op.get("valor")
    color_key = str(op.get("color", "rojo")).lower()
    color_hex = _COLORES_HEX.get(color_key, "FF6B6B")

    if col not in df.columns or valor_ref is None:
        return

    fn_cond = _COMPARADORES.get(condicion)
    if not fn_cond:
        return

    col_idx = list(df.columns).index(col) + 1
    try:
        v_ref = float(valor_ref)
    except (TypeError, ValueError):
        return

    fill_color = PatternFill("solid", fgColor=color_hex)
    for ri in range(2, len(df) + 2):
        celda = ws.cell(row=ri, column=col_idx)
        try:
            if fn_cond(float(celda.value), v_ref):
                celda.fill = fill_color
        except (TypeError, ValueError):
            pass


# ── Nuevas operaciones F1 ────────────────────────────────────────────────────

def _buscar_reemplazar(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Sustituye un valor por otro en una columna o en todo el DataFrame.

    op: buscar (str|num), reemplazar (str|num), col (opcional).
    """
    buscar     = op.get("buscar")
    reemplazar = op.get("reemplazar", "")
    col        = op.get("col")

    if buscar is None:
        raise EditorError("Falta el campo 'buscar'.")

    if col:
        _validar_col(df, col)
        antes = df[col].astype(str).str.contains(str(buscar), na=False).sum()
        df[col] = df[col].replace(buscar, reemplazar)
        # También intentar reemplazo en texto por si el valor es string
        df[col] = df[col].astype(str).str.replace(str(buscar), str(reemplazar), regex=False)
        # Revertir a numérico si procede (errors="ignore" deprecado en pandas futuro)
        try:
            df[col] = pd.to_numeric(df[col])
        except (ValueError, TypeError):
            pass  # la columna contiene texto, se mantiene como str
        desc = f"Reemplazado '{buscar}' por '{reemplazar}' en columna '{col}' ({antes} celdas afectadas)"
    else:
        df = df.replace(buscar, reemplazar)
        df = df.apply(lambda s: s.str.replace(str(buscar), str(reemplazar), regex=False)
                      if s.dtype == object else s)
        desc = f"Reemplazado '{buscar}' por '{reemplazar}' en todo el archivo"

    return df, desc


def _dividir_columna(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Divide una columna de texto en dos o más columnas nuevas.

    op: col, separador (default ' '), col_nueva_1, col_nueva_2, n (nº de partes, default 2).
    """
    col       = op.get("col")
    separador = op.get("separador", " ")
    nombre_1  = op.get("col_nueva_1") or f"{col}_1"
    nombre_2  = op.get("col_nueva_2") or f"{col}_2"
    n_partes  = int(op.get("n", 2))

    if not col:
        raise EditorError("Falta el campo 'col'.")
    _validar_col(df, col)

    partes = df[col].astype(str).str.split(separador, n=n_partes - 1, expand=True)

    nuevas = [nombre_1, nombre_2] + [f"{col}_{i+3}" for i in range(partes.shape[1] - 2)]
    for i, nombre in enumerate(nuevas[:partes.shape[1]]):
        df[nombre] = partes[i] if i < partes.shape[1] else ""

    desc = (f"Columna '{col}' dividida en {partes.shape[1]} columnas "
            f"({', '.join(nuevas[:partes.shape[1]])}) usando '{separador}' como separador")
    return df, desc


def _concatenar_columnas(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Une varias columnas en una nueva columna de texto.

    op: columnas (lista), separador (default ' '), col_resultado.
    """
    columnas     = op.get("columnas", [])
    separador    = op.get("separador", " ")
    col_resultado = op.get("col_resultado") or "_".join(columnas)

    if not columnas or len(columnas) < 2:
        raise EditorError("Se necesitan al menos dos columnas en 'columnas'.")
    for c in columnas:
        _validar_col(df, c)

    df[col_resultado] = df[columnas[0]].astype(str)
    for c in columnas[1:]:
        df[col_resultado] = df[col_resultado] + separador + df[c].astype(str)

    desc = (f"Columnas {', '.join(columnas)} concatenadas en '{col_resultado}' "
            f"con separador '{separador}'")
    return df, desc


# ── Combinar dos DataFrames (B3) ─────────────────────────────────────────────

_TIPOS_JOIN = {"inner", "left", "right", "outer"}
_NOMBRES_JOIN = {
    "inner": "solo filas coincidentes",
    "left":  "todas las filas del primer archivo",
    "right": "todas las filas del segundo archivo",
    "outer": "todas las filas de ambos archivos",
}


def combinar_dataframes(df1: pd.DataFrame, df2: pd.DataFrame,
                        op: dict) -> tuple[pd.DataFrame, str]:
    """Une dos DataFrames por una columna clave.

    op puede contener:
      col  → columna clave común (si es None se elige la primera columna común)
      como → tipo de join: inner / left / right / outer  (default: inner)
    """
    col  = op.get("col")
    como = str(op.get("como", "inner")).lower()
    if como not in _TIPOS_JOIN:
        como = "inner"

    # Si no se especifica columna, usar la primera común
    if not col:
        comunes = [c for c in df1.columns if c in df2.columns]
        if not comunes:
            raise EditorError(
                "No hay columnas en común entre los dos archivos. "
                "Especifica la columna clave."
            )
        col = comunes[0]

    if col not in df1.columns:
        raise EditorError(
            f"Columna '{col}' no existe en el primer archivo. "
            f"Disponibles: {', '.join(df1.columns)}"
        )
    if col not in df2.columns:
        raise EditorError(
            f"Columna '{col}' no existe en el segundo archivo. "
            f"Disponibles: {', '.join(df2.columns)}"
        )

    # Sufijos para columnas duplicadas (excepto la clave)
    cols_dup = [c for c in df1.columns if c in df2.columns and c != col]
    sufijos  = ("_A", "_B") if cols_dup else ("", "")

    try:
        df_result = df1.merge(df2, on=col, how=como, suffixes=sufijos)
    except Exception as exc:
        raise EditorError(f"No se pudo combinar: {exc}") from exc

    desc = (
        f"Archivos combinados por '{col}' ({_NOMBRES_JOIN[como]}). "
        f"Resultado: {len(df_result)} filas × {len(df_result.columns)} columnas"
    )
    return df_result, desc


def _añadir_fila_total(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Añade una fila de totales al final del DataFrame.

    Suma columnas numéricas; pone la etiqueta en la primera columna de texto.
    """
    etiqueta = op.get("etiqueta", "Total")
    funciones = {
        "suma":    lambda s: s.sum(),
        "promedio": lambda s: s.mean(),
        "max":     lambda s: s.max(),
        "min":     lambda s: s.min(),
    }
    aggfunc = funciones.get(op.get("aggfunc", "suma"), funciones["suma"])

    nueva_fila: dict = {}
    primer_texto = True
    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            nueva_fila[col] = aggfunc(pd.to_numeric(df[col], errors="coerce"))
        elif primer_texto:
            nueva_fila[col] = etiqueta
            primer_texto = False
        else:
            nueva_fila[col] = ""

    df = pd.concat([df, pd.DataFrame([nueva_fila])], ignore_index=True)
    return df, f"Fila de {op.get('aggfunc', 'totales')} añadida (etiqueta: '{etiqueta}')"


def _duplicar_filas(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    indices = op.get("indices")
    if indices:
        filas_a_dup = df.iloc[[i for i in indices if 0 <= i < len(df)]]
    else:
        n = max(1, int(op.get("n", 3)))
        filas_a_dup = df.tail(n)
    if filas_a_dup.empty:
        raise EditorError("No hay filas para duplicar.")
    df = pd.concat([df, filas_a_dup], ignore_index=True)
    destino_txt = "al principio" if op.get("destino") == "principio" else "al final"
    return df, f"{len(filas_a_dup)} fila(s) duplicada(s) {destino_txt}"


def _transponer(df: pd.DataFrame, op: dict) -> tuple[pd.DataFrame, str]:
    """Transpone filas y columnas del DataFrame."""
    col_cabecera = op.get("col_cabecera")
    if col_cabecera and col_cabecera in df.columns:
        df_t = df.set_index(col_cabecera).T.reset_index()
        df_t.columns = [str(c) for c in df_t.columns]
        df_t = df_t.rename(columns={"index": col_cabecera})
    else:
        df_t = df.T.reset_index()
        df_t.columns = [str(c) for c in df_t.columns]
    return df_t, f"Tabla transpuesta ({df.shape[1]} columnas → {df_t.shape[1]} columnas)"


# ── Validación ────────────────────────────────────────────────────────────────

def _validar_col(df: pd.DataFrame, col: str | None) -> None:
    if not col:
        raise EditorError("Falta especificar la columna.")
    if col not in df.columns:
        cols = ", ".join(f"'{c}'" for c in df.columns)
        raise EditorError(f"Columna '{col}' no existe. Disponibles: {cols}")
